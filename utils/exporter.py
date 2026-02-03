"""Enhanced data export utilities for various formats - 39 column format."""
import csv
import json
from typing import List, Dict, Optional
from pathlib import Path
from datetime import datetime
from loguru import logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from database import db_manager, BusinessLead
from config.export_columns import (
    EXPORT_COLUMNS,
    FULL_EXPORT_COLUMNS,
    COLD_CALLING_COLUMNS,
    EMAIL_CAMPAIGN_COLUMNS,
    SOCIAL_MEDIA_COLUMNS,
)


class DataExporter:
    """Export scraped data to various formats with enhanced 39-column support."""

    # Use centralized column definitions from config/export_columns.py
    EXPORT_COLUMNS = EXPORT_COLUMNS
    FULL_EXPORT_COLUMNS = FULL_EXPORT_COLUMNS

    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_csv(
        self,
        data: Optional[List[Dict]] = None,
        filters: Optional[Dict] = None,
        filename: Optional[str] = None,
        use_full_format: bool = False
    ) -> str:
        """
        Export data to CSV file in 39-column format.

        Args:
            data: List of business data dictionaries (if None, fetch from DB)
            filters: Database filters to apply when fetching data
            filename: Output filename (auto-generated if None)
            use_full_format: Whether to include all columns (50+) or standard 39

        Returns:
            Path to the exported CSV file
        """
        try:
            if data is None:
                data = self._fetch_from_database(filters)

            if not data:
                logger.warning("No data to export")
                return None

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"leads_export_{timestamp}.csv"

            filepath = self.output_dir / filename

            columns = self.FULL_EXPORT_COLUMNS if use_full_format else self.EXPORT_COLUMNS
            fieldnames = [col[0] for col in columns]

            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()

                for row in data:
                    export_row = self._transform_row(row, columns)
                    writer.writerow(export_row)

            logger.info(f"Exported {len(data)} records to CSV: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise

    def export_to_json(
        self,
        data: Optional[List[Dict]] = None,
        filters: Optional[Dict] = None,
        filename: Optional[str] = None
    ) -> str:
        """Export data to JSON file."""
        try:
            if data is None:
                data = self._fetch_from_database(filters)

            if not data:
                logger.warning("No data to export")
                return None

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"leads_export_{timestamp}.json"

            filepath = self.output_dir / filename

            export_data = {
                'export_date': datetime.now().isoformat(),
                'total_records': len(data),
                'columns': [col[0] for col in self.EXPORT_COLUMNS],
                'leads': []
            }

            for row in data:
                export_row = self._transform_row(row, self.EXPORT_COLUMNS)
                export_data['leads'].append(export_row)

            with open(filepath, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False, default=str)

            logger.info(f"Exported {len(data)} records to JSON: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            raise

    def export_cold_calling_format(
        self,
        filters: Optional[Dict] = None,
        filename: Optional[str] = None
    ) -> str:
        """Export in cold calling optimized format."""
        try:
            if filters is None:
                filters = {}
            filters['has_phone'] = True

            data = self._fetch_from_database(filters)

            if not data:
                logger.warning("No data with phone numbers to export")
                return None

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"cold_calling_leads_{timestamp}.csv"

            filepath = self.output_dir / filename

            # Use centralized cold calling columns
            fieldnames = [col[0] for col in COLD_CALLING_COLUMNS]

            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()

                for row in data:
                    export_row = self._transform_row(row, COLD_CALLING_COLUMNS)
                    writer.writerow(export_row)

            logger.info(f"Exported {len(data)} cold calling leads to: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting cold calling format: {e}")
            raise

    def export_email_campaign_format(
        self,
        filters: Optional[Dict] = None,
        filename: Optional[str] = None
    ) -> str:
        """Export format optimized for email campaigns."""
        try:
            if filters is None:
                filters = {}
            filters['has_email'] = True

            data = self._fetch_from_database(filters)

            if not data:
                logger.warning("No data with emails to export")
                return None

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"email_campaign_leads_{timestamp}.csv"

            filepath = self.output_dir / filename

            # Use centralized email campaign columns
            fieldnames = [col[0] for col in EMAIL_CAMPAIGN_COLUMNS]

            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()

                for row in data:
                    export_row = self._transform_row(row, EMAIL_CAMPAIGN_COLUMNS)
                    writer.writerow(export_row)

            logger.info(f"Exported {len(data)} email campaign leads to: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting email campaign format: {e}")
            raise

    def export_social_media_format(
        self,
        filters: Optional[Dict] = None,
        filename: Optional[str] = None
    ) -> str:
        """Export format with social media links."""
        try:
            if filters is None:
                filters = {}
            filters['has_social'] = True

            data = self._fetch_from_database(filters)

            if not data:
                logger.warning("No data with social media to export")
                return None

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"social_media_leads_{timestamp}.csv"

            filepath = self.output_dir / filename

            # Use centralized social media columns
            fieldnames = [col[0] for col in SOCIAL_MEDIA_COLUMNS]

            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()

                for row in data:
                    export_row = self._transform_row(row, SOCIAL_MEDIA_COLUMNS)
                    writer.writerow(export_row)

            logger.info(f"Exported {len(data)} social media leads to: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting social media format: {e}")
            raise

    def export_to_excel(
        self,
        data: Optional[List[Dict]] = None,
        filters: Optional[Dict] = None,
        filename: Optional[str] = None,
        use_full_format: bool = False
    ) -> str:
        """Export data to Excel (.xlsx) file with professional formatting."""
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")

        try:
            if data is None:
                data = self._fetch_from_database(filters)

            if not data:
                logger.warning("No data to export")
                return None

            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"leads_export_{timestamp}.xlsx"

            if not filename.endswith('.xlsx'):
                filename += '.xlsx'

            filepath = self.output_dir / filename

            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"

            columns = self.FULL_EXPORT_COLUMNS if use_full_format else self.EXPORT_COLUMNS

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Category header fills
            category_fills = {
                'business': PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),
                'location': PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),
                'ratings': PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
                'email': PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
                'phone': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
                'social': PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
            }

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            for col_idx, (export_name, db_field, display_name, width) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=display_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                export_row = self._transform_row(row_data, columns)

                for col_idx, (export_name, _, _, _) in enumerate(columns, 1):
                    value = export_row.get(export_name)

                    # Convert datetime to string
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")

                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border

                    # Highlight quality scores
                    if export_name == 'data_quality_score' and value:
                        if value >= 80:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value >= 60:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif value < 40:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                    # Highlight ratings
                    if export_name == 'rating' and value:
                        if value >= 4.5:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value < 3.0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Add auto-filter
            ws.auto_filter.ref = ws.dimensions

            # Save
            wb.save(filepath)

            logger.info(f"Exported {len(data)} records to Excel: {filepath}")
            return str(filepath)

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise

    def _transform_row(self, row: Dict, columns: List[tuple]) -> Dict:
        """Transform a database row to export format."""
        export_row = {}

        for export_name, db_field, _, _ in columns:
            value = row.get(db_field)

            # Handle special cases
            if db_field == 'subcategories' and value:
                if isinstance(value, list):
                    value = ', '.join(value)

            if db_field == 'scraped_at' and value:
                if isinstance(value, datetime):
                    value = value.isoformat()

            # Handle email fallback (use primary email if email_1 is empty)
            if export_name == 'email_1' and not value:
                value = row.get('email')

            # Handle phone fallback (use primary phone if phone_1 is empty)
            if export_name == 'phone_1' and not value:
                value = row.get('phone')

            export_row[export_name] = value

        return export_row

    def _fetch_from_database(self, filters: Optional[Dict] = None) -> List[Dict]:
        """Fetch data from database with optional filters."""
        try:
            with db_manager.get_session() as session:
                query = session.query(BusinessLead)

                if filters:
                    if filters.get('has_phone'):
                        query = query.filter(BusinessLead.phone.isnot(None))

                    if filters.get('has_website'):
                        query = query.filter(BusinessLead.website.isnot(None))

                    if filters.get('has_email'):
                        query = query.filter(
                            (BusinessLead.email.isnot(None)) |
                            (BusinessLead.email_1.isnot(None))
                        )

                    if filters.get('has_social'):
                        query = query.filter(
                            (BusinessLead.social_facebook.isnot(None)) |
                            (BusinessLead.social_instagram.isnot(None)) |
                            (BusinessLead.social_linkedin.isnot(None)) |
                            (BusinessLead.social_twitter.isnot(None))
                        )

                    if filters.get('city'):
                        query = query.filter(BusinessLead.city == filters['city'])

                    if filters.get('state'):
                        query = query.filter(BusinessLead.state == filters['state'])

                    if filters.get('country'):
                        query = query.filter(BusinessLead.country == filters['country'])

                    if filters.get('category'):
                        query = query.filter(BusinessLead.category == filters['category'])

                    if filters.get('min_quality_score'):
                        query = query.filter(
                            BusinessLead.data_quality_score >= filters['min_quality_score']
                        )

                    if filters.get('max_quality_score'):
                        query = query.filter(
                            BusinessLead.data_quality_score <= filters['max_quality_score']
                        )

                    if filters.get('min_rating'):
                        query = query.filter(BusinessLead.rating >= filters['min_rating'])

                    if filters.get('max_rating'):
                        query = query.filter(BusinessLead.rating <= filters['max_rating'])

                    if filters.get('min_reviews'):
                        query = query.filter(BusinessLead.review_count >= filters['min_reviews'])

                    if filters.get('search_query'):
                        query = query.filter(BusinessLead.search_query == filters['search_query'])

                    if filters.get('founded_after'):
                        query = query.filter(BusinessLead.founded_year >= filters['founded_after'])

                    if filters.get('founded_before'):
                        query = query.filter(BusinessLead.founded_year <= filters['founded_before'])

                    if filters.get('has_employees'):
                        query = query.filter(BusinessLead.employees.isnot(None))

                    if filters.get('has_revenue'):
                        query = query.filter(BusinessLead.revenue.isnot(None))

                    # Limit results
                    if filters.get('limit'):
                        query = query.limit(filters['limit'])

                    # Order by
                    if filters.get('order_by') == 'quality':
                        query = query.order_by(BusinessLead.data_quality_score.desc())
                    elif filters.get('order_by') == 'rating':
                        query = query.order_by(BusinessLead.rating.desc())
                    elif filters.get('order_by') == 'reviews':
                        query = query.order_by(BusinessLead.review_count.desc())
                    elif filters.get('order_by') == 'newest':
                        query = query.order_by(BusinessLead.scraped_at.desc())

                results = query.all()
                data = [lead.to_dict() for lead in results]

                logger.info(f"Fetched {len(data)} records from database")
                return data

        except Exception as e:
            logger.error(f"Error fetching from database: {e}")
            return []

    def get_export_stats(self) -> Dict:
        """Get statistics about exported files."""
        try:
            stats = {
                'total_files': 0,
                'total_size_mb': 0,
                'files': []
            }

            for filepath in self.output_dir.glob('*'):
                if filepath.is_file():
                    size_mb = filepath.stat().st_size / (1024 * 1024)
                    stats['files'].append({
                        'name': filepath.name,
                        'size_mb': round(size_mb, 2),
                        'modified': datetime.fromtimestamp(filepath.stat().st_mtime).isoformat(),
                        'format': filepath.suffix.replace('.', ''),
                    })
                    stats['total_files'] += 1
                    stats['total_size_mb'] += size_mb

            stats['total_size_mb'] = round(stats['total_size_mb'], 2)
            stats['files'] = sorted(stats['files'], key=lambda x: x['modified'], reverse=True)

            return stats

        except Exception as e:
            logger.error(f"Error getting export stats: {e}")
            return {}

    def get_available_formats(self) -> List[Dict]:
        """Get list of available export formats with descriptions."""
        return [
            {
                'id': 'csv',
                'name': 'CSV (Standard)',
                'description': '39 columns - Compatible with Excel, Google Sheets',
                'extension': '.csv',
            },
            {
                'id': 'csv_full',
                'name': 'CSV (Full)',
                'description': '50+ columns - All available data',
                'extension': '.csv',
            },
            {
                'id': 'excel',
                'name': 'Excel',
                'description': 'Formatted Excel file with color coding',
                'extension': '.xlsx',
            },
            {
                'id': 'json',
                'name': 'JSON',
                'description': 'Structured JSON format',
                'extension': '.json',
            },
            {
                'id': 'cold_calling',
                'name': 'Cold Calling',
                'description': 'Optimized for phone outreach',
                'extension': '.csv',
            },
            {
                'id': 'email_campaign',
                'name': 'Email Campaign',
                'description': 'Optimized for email marketing',
                'extension': '.csv',
            },
            {
                'id': 'social_media',
                'name': 'Social Media',
                'description': 'Focus on social media links',
                'extension': '.csv',
            },
        ]


# Convenience functions that delegate to DataExporter class
def export_to_excel(data: List[Dict], filepath: str) -> str:
    """Export data to Excel file using DataExporter."""
    exporter = DataExporter(output_dir=str(Path(filepath).parent))
    return exporter.export_to_excel(data=data, filename=Path(filepath).name)


def export_to_csv(data: List[Dict], filepath: str) -> str:
    """Export data to CSV file using DataExporter."""
    exporter = DataExporter(output_dir=str(Path(filepath).parent))
    return exporter.export_to_csv(data=data, filename=Path(filepath).name)


def export_to_json(data: List[Dict], filepath: str) -> str:
    """Export data to JSON file using DataExporter."""
    exporter = DataExporter(output_dir=str(Path(filepath).parent))
    return exporter.export_to_json(data=data, filename=Path(filepath).name)
